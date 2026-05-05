import tkinter as new
import openpyxl as op
def pasok():
    newenter = entery.get()
    wrk = op.Workbook()
    sheet = wrk.active
    sheet.append([newenter])
    wrk.save("openout.xlsx")
def show():
    lod = op.load_workbook("openout.xlsx")
    shet = lod.active
    for i in shet.iter_cols(values_only=True):
        print(i)
window = new.Tk()
window.title("Simple PRactice")
entery = new.Entry(window)
entery.pack()

btn = new.Button(window,text="enter",command=pasok)
btn.pack()
showbtn = new.Button(window,text="show",command=show)
showbtn.pack()
delbtn = new.Button(window,text="Delete")
delbtn.pack()
window.mainloop()