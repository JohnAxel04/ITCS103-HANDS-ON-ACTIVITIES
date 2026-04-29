import openpyxl as xl 
wrk = xl.Workbook()
sheet = wrk.active
sheet['a1'] = "Fishdog"
sheet['b1'] = "Fishlay"
sheet['c1'] = "Fishwow"
wrk.save("Prac.xlsx")
print("success")

worlk = xl.load_workbook("Prac.xlsx")
shet = worlk.active
for i in shet.iter_rows(values_only=True):
    print(i)