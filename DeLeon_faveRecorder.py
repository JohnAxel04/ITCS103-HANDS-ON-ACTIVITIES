import openpyxl as op
import time
import os
os.system("cls")
list = "First","Second","Third"
id = 0
lists = "ID","First Name","Last Name","Birth Date","Age"
work = op.Workbook()
sheet = work.active
sheet.append(lists)
while True:
    for i in list:
        l3 = (f"{i} Favorite Person\n\n")
        for text in l3:
            print(text,end="",flush=True)
            time.sleep(.03)
        fName = input("First Name: ")
        lName = input("Last Name: ")
        bYear = input("Birth Year: ")
        print("")
        if not bYear.isdigit():
            os.system("cls")
            l4  = "Birth Years must be number"
            for text in l4:
                print(text,end="",flush=True)
                time.sleep(.03)
            time.sleep(.3)
            l5 = "\nInput not saved\n"
            for text in l5:
                print(text,end="",flush=True)
                time.sleep(.03)
            time.sleep(.3)
            l5s = "Run the system again\n"
            for text in l5s:
                print(text,end="",flush=True)
                time.sleep(.03)
            time.sleep(.3)
            l5ss = "Thank you!\n"
            for text in l5ss:
                print(text,end="",flush=True)
                time.sleep(.03)
            time.sleep(.3)
            exit()
            
        
        id += 1
        newb = int(bYear)
        age = 2026 - newb
        
        sheet.append([f"0{id}",fName,lName,newb,age])
        work.save("favorite_people.xlsx")
        os.system("cls")
    
    
    l7 = "File saving..."
    for text in l7:
        print(text,end="",flush=True)
        time.sleep(.2)
    time.sleep(.2)
    os.system("cls")
    l8 = "..."
    for text in l8:
        print(text,flush=True)
        time.sleep(.3)
    time.sleep(.2)
    os.system("cls")
    l2 = ("\nFavorite People List\n\n")
    for text in l2:
        print(text,end="",flush=True)
        time.sleep(.03)
    lwork = op.load_workbook("favorite_people.xlsx")
    lsheets = lwork.active
    for i in lsheets.iter_rows(values_only=True):
        print(i)
    ask = input("\nPress Enter to Exit... ")
    os.system("cls")
    
    if ask == "":
        l1 = ("\nThank you for using my system")
        for text in l1:
            print(text,end="",flush=True)
            time.sleep(.03)
        break
    else:
        continue
    