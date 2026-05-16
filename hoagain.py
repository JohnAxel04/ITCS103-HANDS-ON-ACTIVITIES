import tkinter as new
import openpyxl as op
from tkinter import ttk,messagebox

work = op.Workbook()
sheet = work.active
sheet["a1"] = "ID"
sheet["b1"] = "Full Name"
sheet["c1"] = "Birth Date"
sheet["d1"] = "Age"
sheet["e1"] = "Gender"
work.save("letsgo.xlsx")


def inputvalidation():
    full = nameEntry.get()
    birth = birthEntry.get()
    
    if not full or not birth:
        messagebox.showerror("Input Invalid","Entry must not be empty")
        return False
    if not birth.isdigit():
        messagebox.showerror("Input Invalid","Birth date must be a number")
        return False
    
    return True
    
def save():
    if not inputvalidation():
        return
    
    full = nameEntry.get()
    birth = int(birthEntry.get())
    age = 2026 - birth
    gender = var.get()

    
    nameEntry.delete(0,new.END)
    birthEntry.delete(0,new.END)

    loads = op.load_workbook("letsgo.xlsx")
    sheets = loads.active
    id = sheets.max_row
    sheets.append([id,full,birth,age,gender])
    loads.save("letsgo.xlsx")
    show()
    messagebox.showinfo("Success","You file is successfully saved")

def show():
    workload = op.load_workbook("letsgo.xlsx")
    sht = workload.active
    for i in table.get_children():
        table.delete(i)

    for i in sht.iter_rows(min_row=2,values_only=True):
        table.insert("",new.END,values=i)

def select(event):
    selects = table.focus()
    values = table.item(selects,"values")

    if values:
        nameEntry.delete(0,new.END)
        birthEntry.delete(0,new.END)

        nameEntry.insert(0,values[1])
        birthEntry.insert(0,values[2])

def update():
    selected = table.focus()
    values = table.item(selected,"values")

    if not selected:
        messagebox.showerror("Not selected","Select first to update data")
        return

    if not inputvalidation:
        return
    
    lw = op.load_workbook("letsgo.xlsx")
    shat = lw.active

    ids = values[0]
    name = nameEntry.get()
    bth = int(birthEntry.get())
    age = 2026 - bth

    for row in shat.iter_rows(min_row=2):
        if int(row[0].value) == int(ids):
            row[1].value = name
            row[2].value = bth
            row[3].value = age
    lw.save("letsgo.xlsx")

    messagebox.showinfo("Update Successfull","Data updated sucessfully")
    show()

def deletes():
    sel = table.focus
    values = table.item(sel,"values")
    record_id = values[0]

    if not sel:
        messagebox.showerror("Input Invalid","Select first to delete data")
        return
    
    confirm = messagebox.askyesnocancel("Confirmation","Are you sure u want to delete?")
    if not confirm:
        return

    lwrk = op.load_workbook("letsgo.xlsx")
    sht = lwrk.active

    for i,row in enumerate(sht.iter_rows(min_row=2),start=2):
        if int(row[0].value) == int(record_id):
            sht.delete_rows(i)
            break

    lwrk.save("letsgo.xlsx")
    messagebox.showinfo("Delete","Data successfully Deleted")
    show()
    
        
    

window = new.Tk()
window.title("Simple User Input")
titles = new.Label(window,text="User Input")
titles.grid(column=0,columnspan=3,row=0)
namelabel = new.Label(window,text="Full Name:")
namelabel.grid(column=0,row=1)
nameEntry = new.Entry(window)
nameEntry.grid(column=1,columnspan=2,row=1)

var = new.StringVar()
var.set("Male")
male = new.Radiobutton(window,text="Male",value="Male",variable=var)
male.grid(column=0,row=3)
female = new.Radiobutton(window,text="Female",value="Female",variable=var)
female.grid(column=1,row=3)


birthlabel = new.Label(window,text="BirthYear:")
birthlabel.grid(column=0,row=2)
birthEntry = new.Entry(window)
birthEntry.grid(row=2,column=1,columnspan=2)
updatebtn = new.Button(window,text="Update",command=update)
updatebtn.grid(column=0,row=4)
savebtn = new.Button(window,text="Save",command=save)
savebtn.grid(column=1,row=4)
delbtn = new.Button(window,text="Delete",command=deletes)
delbtn.grid(column=2,row=4)

table = ttk.Treeview(window,columns=("ID","Full Name","Birth Year","Age","Gender"),show="headings")
for col in ("ID","Full Name","Birth Year","Age","Gender"):
    table.heading(col,text=col)
table.grid(row=5,pady=10,columnspan=3)

table.bind("<<TreeviewSelect>>",select)

show()
window.mainloop()
