import tkinter as new
import openpyxl as op
def submitted():
    newname = nameEntry.get()
    newage = ageEntry.get()
    newcourse = courseEntry.get()

    wrk = op.Workbook()
    sheet = wrk.active
    sheet.append([newname,newage,newcourse])
    wrk.save("File.xlsx")
def view():
    top = new.Toplevel(window)
    top.transient(window)
    top.grab_set()
    toplabel = new.Label(top,text="View Saved Files")
    toplabel.grid(columnspan=3)
    lbl = new.Label(top,text="Name Age Course")
    lbl.grid(row=1,columnspan=3)
    labelview = new.Label(top,width=10)
    labelview.grid(row=2,columnspan=3)
    wrkk = op.load_workbook("File.xlsx")
    sheets = wrkk.active
    for i in sheets.iter_rows(values_only=True):
        labelview['text'] = i
    
    
window = new.Tk()
window.title("Openpyxl")
mainlabel = new.Label(window,text="Fill up the blank")
mainlabel.grid(columnspan=3)
namelabel = new.Label(window,text="name: ")
namelabel.grid(row=2)
nameEntry = new.Entry(window)
nameEntry.grid(row=2,column=1,columnspan=2)
agelabel = new.Label(window,text="age: ")
agelabel.grid(row=3,column=0)
ageEntry = new.Entry(window)
ageEntry.grid(row=3,column=1,columnspan=2)
courseLabel = new.Label(window,text="Course: ")
courseLabel.grid(row=4,column=0)
courseEntry = new.Entry(window)
courseEntry.grid(row=4,column=1,columnspan=2)
btn = new.Button(window,text="Submit",command=submitted,width=10)
btn.grid(row=5,column=0,columnspan=4)
readbtn = new.Button(window,text="View",command=view,width=10)
readbtn.grid(row=6,columnspan=4)
window.mainloop()