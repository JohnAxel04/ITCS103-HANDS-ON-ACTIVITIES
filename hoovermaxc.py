import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

workbook = op.Workbook()
sht = workbook.active
sht['a1'] = "ID"
sht['b1'] = "Last"
sht['c1'] = "First"
sht['d1'] = "Middle"
sht['e1'] = "Birthdate"
sht['f1'] = "Age"
workbook.save("workmoto.xlsx")
def checker():
    last = lname_entry.get()
    first = fname_entry.get()
    mid = mname_entry.get()
    births = birth_entry.get()

    if not last or not first or not mid or not births:
        messagebox.showinfo("error","must not be empty")
        return False
    if not births.isdigit():
        messagebox.showinfo("error","must be a number")
        return False
    return True
def save():
    if not checker():
        return
    last = lname_entry.get()
    first = fname_entry.get()
    mid = mname_entry.get()
    births = int(birth_entry.get())
    age = 2026 - births

    loads = op.load_workbook("workmoto.xlsx")

    shitt = loads.active

    newid = shitt.max_row

    shitt.append([newid,last,first,mid,births,age])

    loads.save("workmoto.xlsx")
    messagebox.showinfo("ewrwer","workmoto.xlsx")
    show()
    print("success")

def show():
    ld = op.load_workbook("workmoto.xlsx")
    shut = ld.active
    for i in table.get_children():
        table.delete(i)
    for q in shut.iter_rows(min_row=2,values_only=True):
        table.insert("",tk.END,values=q)

def focuss(event):
    selected = table.focus()
    values = table.item(selected,"values")

    if values:
        fname_entry.delete(0,tk.END)
        lname_entry.delete(0,tk.END)
        mname_entry.delete(0,tk.END)
        birth_entry.delete(0,tk.END)

        fname_entry.insert(0,values[2])
        lname_entry.insert(0,values[1])
        mname_entry.insert(0,values[3])
        birth_entry.insert(0,values[4])

def update():
    select = table.focus()
    values = table.item(select,"values")
    if not select:
        messagebox.showerror("wq","werwqe")
    if not checker():
        messagebox.showerror("wqweqrqwerqwe","werrwwqe")
    first = fname_entry.get()
    mid = mname_entry.get()
    last = lname_entry.get()
    birth = int(birth_entry.get())
    age = 26 - birth
    new_id = values[0]
    loader = op.load_workbook("workmoto.xlsx")
    shut = loader.active
    for i in shut.iter_rows(min_row=2):
        if str(i[0].value) == str(new_id):
            i[1].value = last
            i[2].value = first
            i[3].value = mid
            i[4].value = birth
            i[5].value = age
    loader.save("workmoto.xlsx")
    messagebox.showinfo("suc","sycsadf")
    show()

def delete():
    selectmoto = table.focus()
    valueto = table.item(selectmoto,"values")

    if not selectmoto:
        messagebox.showerror("error","wenrjwqer")

    maxid = valueto[0]

    confirm = messagebox.askyesnocancel("werwe","wqerwqe")
    if not confirm:
        return
    
    loads = op.load_workbook("workmoto.xlsx")
    shyt = loads.active
    for i,row in enumerate(shyt.iter_rows(min_row=2),start=2):
        if str(row[0].value) == str(maxid):
            shyt.delete_rows(i)
    loads.save("workmoto.xlsx")
    messagebox.showinfo("werhiwerw","wrwerwe")
    show()



window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

update_btn = tk.Button(window, text="Update",command=update)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=save)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white",command=delete)
delete_btn.grid(row=6, column=3)

table = ttk.Treeview(window,columns=("ID","Last Name","First Name","Middle Name","Birth Date","Age"), show="headings")
for col in ("ID","Last Name","First Name","Middle Name","Birth Date","Age"):
    table.heading(col,text=col)
table.grid(row=7, columnspan=4)
table.bind("<<TreeviewSelect>>",focuss)
show()
window.mainloop()