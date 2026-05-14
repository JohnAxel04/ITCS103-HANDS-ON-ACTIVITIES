import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

wrk = op.Workbook()
sheet = wrk.active
sheet["a1"] = "ID"
sheet["b1"] = "Last"
sheet["c1"] = "First"
sheet["d1"] = "Middle"
sheet["e1"] = "Birth Year"
sheet["f1"] = "Age"
wrk.save("myFile.xlsx")

def display():
    lwork = op.load_workbook("myFile.xlsx")
    shtt = lwork.active

    for i in table.get_children():
        table.delete(i)

    for i in shtt.iter_rows(values_only=True,min_row=2):
        table.insert("",tk.END,values=i)
    



def validate():
    first = fname_entry.get()
    last = lname_entry.get()
    mid = mname_entry.get()
    birth = birth_entry.get()

    if not first or not last or not mid or not birth:
        messagebox.showerror("Input Invalid","Entry must not be empty")
        return False
    if not birth.isdigit():
        messagebox.showerror("Input Invalid","Birth Year must be a number")
        return False
    
    return True

def save():
    if not validate():
        return
    
    first = fname_entry.get()
    last = lname_entry.get()
    mid = mname_entry.get()
    birth = int(birth_entry.get())
    age = 2026 - birth

    load = op.load_workbook("myFile.xlsx")
    sht = load.active

    newid = sht.max_row

    sht.append([newid,last,first,mid,birth,age])

    load.save("myFile.xlsx")

    messagebox.showinfo("Success","You file is successfully saved")
    display()

def select(events):
    selected = table.focus()
    values = table.item(selected,"values")

    if values:
        lname_entry.delete(0,tk.END)
        fname_entry.delete(0,tk.END)
        mname_entry.delete(0,tk.END)
        birth_entry.delete(0,tk.END)

        lname_entry.insert(0,values[1])
        fname_entry.insert(0,values[2])
        mname_entry.insert(0,values[3])
        birth_entry.insert(0,values[4])

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Input Invalid","Select first")
    
    if not validate():
        return
    
    values = table.item(selected,"values")
    recordid = values[0]

    first = fname_entry.get()
    last = lname_entry.get()
    middle = mname_entry.get()
    bd = int(birth_entry.get())
    age = 2026 - bd

    lwrk = op.load_workbook("myFile.xlsx")
    sheets = lwrk.active

    for i in sheets.iter_rows(min_row=2):
        if str(i[0].value) == str(recordid):
            i[1].value = last
            i[2].value = first
            i[3].value = middle
            i[4].value = bd
            i[5].value = age

    lwrk.save("myFile.xlsx")
    messagebox.showinfo("Sucess","Info Updated Successfully")
    display()

def delete():
    selecte = table.focus()

    if not selecte:
       messagebox.showerror("Input Invalid","Select first")

    values = table.item(selecte,"values")
    ids = values[0]

    confirm = messagebox.askyesnocancel("Delete","Are you sure you want to delete?")
    if not confirm:
        return
    
    wrks = op.load_workbook("myFile.xlsx")
    shat = wrks.active

    for i,row in enumerate(shat.iter_rows(min_row=2,values_only=True),start=2):
        if str(row[0]) == str(ids):
            shat.delete_rows(i)
            break

    wrks.save("myFile.xlsx")
    messagebox.showinfo("Successfully","Input sucessfuly deleted")
    display()
window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

update_btn = tk.Button(window, text="Update",command=update)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=save)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white",command=delete)
delete_btn.grid(row=6, column=3)

table = ttk.Treeview(window, columns=("ID","Last","First","Middle","Birth Year","Age"),show="headings")
for col in ("ID","Last","First","Middle","Birth Year","Age"):
    table.heading(col,text=col)
table.grid(row=7,columnspan=4)

table.bind("<<TreeviewSelect>>", select)
display()
window.mainloop()