import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

work = op.Workbook()
sht = work.active
sht['a1'] = "ID"
sht['b1'] = "Last"
sht['c1'] = "First"
sht['d1'] = "Middle"
sht['e1'] = "Birth Date"
sht['f1'] = "Age"
work.save("excelDb.xlsx")

def display():
    wrkk = op.load_workbook("excelDb.xlsx")
    shet = wrkk.active
    
    for row in tree.get_children():
        tree.delete(row)
    for row in shet.iter_rows(min_row=2,values_only=True):
        tree.insert("",tk.END,values=row)

def inputvalidation():
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if not first or not middle or not last:
        messagebox.showerror("Input Error","All input must not be empty")
        return False
    
    if not birth.isdigit():
        messagebox.showerror("Input Invalid","Birth Year must be a number")
        return False
    return True
def save():
    if not inputvalidation():
        return False
    
    firsts = fname_entry.get()
    middles = mname_entry.get()
    lasts = lname_entry.get()
    births = int(birth_entry.get())
    

    age = 2026 - births

    wrk = op.load_workbook("excelDb.xlsx")
    sheet = wrk.active

    newId = sheet.max_row

    sheet.append([newId,firsts,middles,lasts,births,age])
    wrk.save("excelDb.xlsx")

    display()
    messagebox.showinfo("Success","File successfully saved")

def select(event):
    selection = tree.focus()
    values = tree.item(selection, "values")

    if values:
        lname_entry.delete(0,tk.END)
        fname_entry.delete(0,tk.END)
        mname_entry.delete(0,tk.END)
        birth_entry.delete(0,tk.END)

        lname_entry.insert(0,values[1])
        fname_entry.insert(0,values[2])
        mname_entry.insert(0,values[3])
        birth_entry.insert(0,values[4])

def updt():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error","Select a Record first")
        return
    
    if not inputvalidation():
        return
    
    values = tree.item(selected,"values")
    recordId = values[0]

    first = fname_entry.get()
    midle = mname_entry.get()
    last = lname_entry.get()
    by = int(birth_entry.get())
    age = 2026 - by

    wrkbook = op.load_workbook("excelDb.xlsx")
    act = wrkbook.active

    for row in act.iter_rows(min_row=2):
        if int(row[0].value) == int(recordId):
            row[1].value = last
            row[2].value = first
            row[3].value = midle
            row[4].value = by
            row[5].value = age
    
    wrkbook.save("excelDb.xlsx")

    messagebox.showinfo("Success","Record updated successfully")

    display()
def delete():
    selectedd = tree.focus()

    if not selectedd:
        messagebox.showerror("Error","Select a Record first")
        return

    value = tree.item(selectedd,"values")
    recordId = value[0]

    confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete?")
    if not confirm:
        return

    workb = op.load_workbook("excelDb.xlsx")
    sht = workb.active
    for i,row in enumerate(sht.iter_rows(min_row=2),start=2):
        if int(row[0].value) == int(recordId):
            sht.delete_rows(i)
            break
    
    workb.save("excelDb.xlsx")

    messagebox.showinfo("Success","Record deleted successfully")
    display()
window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

update_btn = tk.Button(window, text="Update",command=updt)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=save)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white",command=delete)
delete_btn.grid(row=6, column=3)

tree = ttk.Treeview(window, columns=("ID","Last","First","Middle","BirthYear","Age"), show="headings")
for col in ("ID","Last","First","Middle","BirthYear","Age"):
    tree.heading(col, text=col)
tree.grid(row=7, column=0, columnspan=4)

tree.bind("<<TreeviewSelect>>", select)

display()
window.mainloop()