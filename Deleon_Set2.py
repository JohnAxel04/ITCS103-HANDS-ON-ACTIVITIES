import tkinter as tk
from tkinter import ttk,messagebox
import openpyxl as op
work = op.Workbook()
sheet = work.active
sheet["a1"] = "Order ID"
sheet["b1"] = "Customer name"
sheet["c1"] = "Product"
sheet["d1"] = "Quantity"
sheet["e1"] = "Price"
sheet["f1"] = "Total"
work.save("ordersDB.xlsx")

def input_validation():
    name = cname_entry.get()
    product = product_entry.get()
    qty = qty_entry.get()
    price = price_entry.get()

    if not name or not product or not qty or not price:
        messagebox.showerror("Invalid Input","Input must not be empty")
        return False
    if not qty.isdigit() or not price.isdigit():
        messagebox.showerror("Invalid Input","Quantity and Price must be a Number")
        return False
    return True

def save():
    if not input_validation():
        return
    
    name = cname_entry.get()
    product = product_entry.get()
    qty = int(qty_entry.get())
    price =int(price_entry.get()) 
    total = qty * price

    load = op.load_workbook("ordersDB.xlsx")
    sheet = load.active
    product_id = sheet.max_row
    sheet.append([product_id,name,product,qty,price,total])
    load.save("ordersDB.xlsx")
    messagebox.showinfo("Successfull","File successfully created")
    show()
def populate(event):
    select = table.focus()
    values = table.item(select,"values")

    if values:
        cname_entry.delete(0,tk.END)
        product_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        price_entry.delete(0,tk.END)

        cname_entry.insert(0,values[1])
        product_entry.insert(0,values[2])
        qty_entry.insert(0,values[3])
        price_entry.insert(0,values[4])
def show():
    load = op.load_workbook("ordersDB.xlsx")
    sheet = load.active
    for i in table.get_children():
        table.delete(i)
    for q in sheet.iter_rows(values_only=True,min_row=2):
        table.insert("",tk.END,values=q)
def delete():
    select = table.focus()
    values = table.item(select,"values")

    if not select:
        messagebox.showerror("Invalid","Select first to delete")
        return
    
    confirm = messagebox.askyesnocancel("Delete","Want to delete?")
    if not confirm:
        return
    new_id = values[0]
    load = op.load_workbook("ordersDB.xlsx")
    sheet = load.active
    for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
        if str(row[0].value) == str(new_id):
            sheet.delete_rows(i)
    load.save("ordersDB.xlsx")
    messagebox.showinfo("Success","File successfully Deleted")
    show()
window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")

# Form Title
title = tk.Label(window, text="Simple Ordering System", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue")
cname_label.grid(row=3, column=1, columnspan=2)

# Product Entry
product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

product_label = tk.Label(genframe, text="Product", font=("Poppins", 10, "italic"), bg="lightblue")
product_label.grid(row=3, column=3, columnspan=2)

# Quantity Entry
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe, text="Quantity", font=("Poppins", 10, "italic"), bg="lightblue")
qty_label.grid(row=5, column=1, columnspan=2)

# Price Entry
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe, text="Price", font=("Poppins", 10, "italic"), bg="lightblue")
price_label.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink", command=save)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen")
update_btn.grid(row=6, column=2)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"),command=delete)
delete_btn.grid(row=6, column=3)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)
table.bind("<<TreeviewSelect>>",populate)
show()
window.mainloop()